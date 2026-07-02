"""Benchmark đa model: chạy bộ câu hỏi qua API với từng LLM → Excel so sánh.

    python -m scripts.benchmark_suite --limit 2            # smoke: 2 câu × N model
    python -m scripts.benchmark_suite --category GT,CALC   # vài nhóm trước
    python -m scripts.benchmark_suite                      # full 77 câu × 6 model (4-8 GIỜ!)
    python -m scripts.benchmark_suite --report-only        # chỉ dựng lại Excel từ kết quả đã có

Resume: kết quả ghi từng dòng vào data/benchmark_results.jsonl — chạy lại sẽ
tự BỎ QUA các cặp (model, câu) đã chạy xong. Dùng --fresh để xóa làm lại.

Latency đo từ lúc gửi câu hỏi đến lúc nhận đủ câu trả lời (stream=False) —
tương đương thời gian người dùng chờ trên UI.
"""
from __future__ import annotations

import argparse
import json
import statistics
import sys
import time
from pathlib import Path

for _s in (sys.stdout, sys.stderr):
    if hasattr(_s, "reconfigure"):
        try:
            _s.reconfigure(encoding="utf-8", errors="replace")
        except Exception:
            pass

import requests

ROOT = Path(__file__).resolve().parent.parent
DEFAULT_FILE = ROOT / "tests" / "demo_questions.json"
DEFAULT_RESULTS = ROOT / "data" / "benchmark_results.jsonl"
DEFAULT_XLSX = ROOT / "data" / "benchmark_report.xlsx"

# "default" = LLM_MODEL trong .env (không gửi override)
DEFAULT_MODELS = ",".join([
    "default",
    "cc/claude-sonnet-4-5-20250929",
    "cc/claude-haiku-4-5-20251001",
    "gh/gpt-4.1",
    "gh/gemini-2.5-pro",
    "kc/deepseek/deepseek-chat",
])

# Thang điểm Latency (TB giây/câu → điểm 1-5)
LATENCY_SCALE = [(20, 5), (40, 4), (60, 3), (90, 2)]

# (tên cột, trọng số, auto?) — tổng trọng số = 100%
CRITERIA = [
    ("Điểm Latency",   0.05, True),
    ("Correctness",    0.20, False),
    ("Hallucination",  0.15, False),
    ("Groundedness",   0.15, False),
    ("Citation",       0.15, False),
    ("Legal reasoning", 0.15, False),
    ("Vietnamese",     0.10, False),
    ("Long-context",   0.05, False),
]


def latency_score(avg_secs: float) -> int:
    for limit, score in LATENCY_SCALE:
        if avg_secs <= limit:
            return score
    return 1


def ask(base_url: str, rag_mode: str, model: str, q: dict, timeout: int) -> tuple[str, float]:
    payload: dict = {
        "model": rag_mode,
        "messages": q.get("history", []) + [{"role": "user", "content": q["question"]}],
        "stream": False,
    }
    if model != "default":
        payload["llm_model"] = model
    t0 = time.time()
    r = requests.post(f"{base_url}/v1/chat/completions", json=payload, timeout=timeout)
    dt = time.time() - t0
    r.raise_for_status()
    return r.json()["choices"][0]["message"]["content"], dt


# ── Excel ─────────────────────────────────────────────────────────────────────

def build_excel(rows: list[dict], questions: list[dict], out: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    bold = Font(bold=True)
    head_fill = PatternFill("solid", fgColor="1F4E5F")
    head_font = Font(bold=True, color="FFFFFF")
    wrap = Alignment(wrap_text=True, vertical="top")

    models = sorted({r["model"] for r in rows})
    by_model: dict[str, list[dict]] = {m: [r for r in rows if r["model"] == m] for m in models}
    q_order = [q["id"] for q in questions]
    q_by_id = {q["id"]: q for q in questions}

    wb = Workbook()

    # ── Sheet 1: Tổng quan ────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Tổng quan"
    headers = ["Model", "Số câu", "PASS", "PASS %", "Lỗi", "TB giây", "P50 giây", "P95 giây"]
    headers += [f"{name} ({int(w*100)}%)" for name, w, _ in CRITERIA]
    headers += ["TỔNG ĐIỂM /5"]
    ws.append(headers)
    first_crit_col = 9  # cột I
    for r_i, m in enumerate(models, start=2):
        rs = by_model[m]
        secs = sorted(x["secs"] for x in rs if x["status"] != "ERROR") or [0]
        n_pass = sum(1 for x in rs if x["status"] == "PASS")
        n_err = sum(1 for x in rs if x["status"] == "ERROR")
        avg = sum(secs) / len(secs)
        row = [m, len(rs), n_pass, round(100 * n_pass / max(len(rs), 1), 1), n_err,
               round(avg, 1), round(statistics.median(secs), 1),
               round(secs[max(0, int(len(secs) * 0.95) - 1)], 1)]
        for c_i, (name, w, auto) in enumerate(CRITERIA):
            row.append(latency_score(avg) if auto else None)  # None = chấm tay
        ws.append(row)
        # Công thức tổng có trọng số
        parts = [
            f"{get_column_letter(first_crit_col + i)}{r_i}*{w}"
            for i, (_, w, _) in enumerate(CRITERIA)
        ]
        ws.cell(row=r_i, column=first_crit_col + len(CRITERIA),
                value="=" + "+".join(parts))
    for c in range(1, len(headers) + 1):
        ws.cell(row=1, column=c).fill = head_fill
        ws.cell(row=1, column=c).font = head_font
        ws.column_dimensions[get_column_letter(c)].width = 14
    ws.column_dimensions["A"].width = 32
    ws.freeze_panes = "B2"

    # ── Sheet 2: Chi tiết (wide — mỗi model 2 cột) ────────────────────────────
    ws2 = wb.create_sheet("Chi tiết")
    h = ["ID", "Nhóm", "Câu hỏi"]
    for m in models:
        h += [f"{m} | KQ", f"{m} | giây"]
    ws2.append(h)
    idx = {(r["model"], r["id"]): r for r in rows}
    for qid in q_order:
        if not any((m, qid) in idx for m in models):
            continue
        q = q_by_id[qid]
        line: list = [qid, q["category"], q["question"]]
        for m in models:
            r = idx.get((m, qid))
            line += [r["status"] if r else "", r["secs"] if r else ""]
        ws2.append(line)
    for c in range(1, len(h) + 1):
        ws2.cell(row=1, column=c).fill = head_fill
        ws2.cell(row=1, column=c).font = head_font
    ws2.column_dimensions["C"].width = 50
    ws2.freeze_panes = "D2"

    # ── Sheet 3: Trả lời (long — để chấm tay) ─────────────────────────────────
    ws3 = wb.create_sheet("Trả lời")
    ws3.append(["Model", "ID", "Nhóm", "Câu hỏi", "Giây", "KQ", "Trả lời đầy đủ"])
    for m in models:
        for r in sorted(by_model[m], key=lambda x: q_order.index(x["id"]) if x["id"] in q_order else 999):
            ws3.append([m, r["id"], r.get("category", ""), r.get("question", ""),
                        r["secs"], r["status"],
                        (r.get("answer") or r.get("error") or "")[:30000]])
    for c in range(1, 8):
        ws3.cell(row=1, column=c).fill = head_fill
        ws3.cell(row=1, column=c).font = head_font
    ws3.column_dimensions["D"].width = 40
    ws3.column_dimensions["G"].width = 90
    for row_cells in ws3.iter_rows(min_row=2, min_col=7, max_col=7):
        for cell in row_cells:
            cell.alignment = wrap
    ws3.freeze_panes = "A2"

    # ── Sheet 4: Thang điểm ───────────────────────────────────────────────────
    ws4 = wb.create_sheet("Thang điểm")
    ws4.append(["Tiêu chí", "Trọng số", "Cách chấm (1-5)"])
    rubric = {
        "Điểm Latency":   "TỰ ĐỘNG từ TB giây/câu: ≤20s=5 | ≤40s=4 | ≤60s=3 | ≤90s=2 | >90s=1",
        "Correctness":    "Đáp án pháp lý đúng (đối chiếu văn bản gốc): 5=đúng hoàn toàn, 3=đúng một phần/thiếu, 1=sai",
        "Hallucination":  "5=không bịa fact/số tiền/điều luật nào; 3=có chi tiết không kiểm chứng được; 1=bịa rõ ràng",
        "Groundedness":   "Khẳng định bám vào căn cứ được trích dẫn trong context: 5=bám hoàn toàn, 1=nói ngoài căn cứ",
        "Citation":       "Số hiệu văn bản + Điều/Khoản đúng, văn bản còn hiệu lực: 5=chính xác, 3=đúng văn bản sai khoản, 1=sai văn bản",
        "Legal reasoning": "Nhóm CALC/FU/CMP: lập luận từng bước, cộng dồn/so sánh đúng logic: 5=chặt chẽ, 1=rối/sai logic",
        "Vietnamese":     "Tiếng Việt tự nhiên, đúng thuật ngữ pháp lý: 5=như luật sư viết, 1=ngô nghê/dịch máy",
        "Long-context":   "Chất lượng khi câu hỏi nhiều dữ kiện hoặc top_k cao: 5=không bỏ sót dữ kiện, 1=lạc đề",
    }
    for name, w, auto in CRITERIA:
        ws4.append([name + (" (tự động)" if auto else ""), f"{int(w*100)}%", rubric[name]])
    ws4.append([])
    ws4.append(["Ghi chú:", "", "PASS/MISS trong các sheet chỉ là kiểm tra TỪ KHÓA tự động (proxy thô) — "
                "điểm Correctness/Hallucination... cần người chấm trên sheet 'Trả lời'. "
                "Điền điểm 1-5 vào sheet 'Tổng quan', cột TỔNG ĐIỂM tự tính theo trọng số."])
    for c in range(1, 4):
        ws4.cell(row=1, column=c).fill = head_fill
        ws4.cell(row=1, column=c).font = head_font
    ws4.column_dimensions["A"].width = 24
    ws4.column_dimensions["C"].width = 100
    for row_cells in ws4.iter_rows(min_row=2, min_col=3, max_col=3):
        for cell in row_cells:
            cell.alignment = wrap

    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)


# ── Main ──────────────────────────────────────────────────────────────────────

def main() -> None:
    p = argparse.ArgumentParser(description="Benchmark đa model qua API")
    p.add_argument("--models", default=DEFAULT_MODELS,
                   help='CSV model ids; "default" = LLM trong .env')
    p.add_argument("--file", default=str(DEFAULT_FILE))
    p.add_argument("--base-url", default="http://127.0.0.1:8000")
    p.add_argument("--rag-mode", default="legal-ai-graph")
    p.add_argument("--category", default=None, help="Lọc nhóm, vd: GT,CALC")
    p.add_argument("--limit", type=int, default=None, help="N câu đầu mỗi model")
    p.add_argument("--delay", type=float, default=2.0)
    p.add_argument("--timeout", type=int, default=300)
    p.add_argument("--results", default=str(DEFAULT_RESULTS))
    p.add_argument("--out", default=str(DEFAULT_XLSX))
    p.add_argument("--fresh", action="store_true", help="Xóa kết quả cũ, chạy lại từ đầu")
    p.add_argument("--report-only", action="store_true", help="Chỉ dựng Excel từ jsonl đã có")
    args = p.parse_args()

    data = json.loads(Path(args.file).read_text(encoding="utf-8"))
    questions = data["questions"]
    if args.category:
        cats = {c.strip().upper() for c in args.category.split(",")}
        questions = [q for q in questions if q["category"].upper() in cats]
    if args.limit:
        questions = questions[: args.limit]

    results_path = Path(args.results)
    if args.fresh and results_path.exists():
        results_path.unlink()
    rows: list[dict] = []
    if results_path.exists():
        for line in results_path.read_text(encoding="utf-8").splitlines():
            if line.strip():
                rows.append(json.loads(line))
    done = {(r["model"], r["id"]) for r in rows}

    if not args.report_only:
        models = [m.strip() for m in args.models.split(",") if m.strip()]
        health = requests.get(f"{args.base_url}/", timeout=10).json()
        todo_total = sum(1 for m in models for q in questions if (m, q["id"]) not in done)
        print(f"API OK ({health.get('chunks', 0):,} chunks) | {len(models)} model × "
              f"{len(questions)} câu | còn phải chạy: {todo_total} call "
              f"(đã có sẵn {len(done)})")

        results_path.parent.mkdir(parents=True, exist_ok=True)
        f = results_path.open("a", encoding="utf-8")
        n_run = 0
        t0 = time.time()
        for m in models:
            for q in questions:
                if (m, q["id"]) in done:
                    continue
                try:
                    answer, dt = ask(args.base_url, args.rag_mode, m, q, args.timeout)
                    low = answer.lower()
                    hits = [k for k in q.get("expect_any", []) if k.lower() in low]
                    rec = {"model": m, "id": q["id"], "category": q["category"],
                           "question": q["question"], "status": "PASS" if hits else "MISS",
                           "secs": round(dt, 1), "answer": answer, "error": ""}
                except Exception as exc:
                    rec = {"model": m, "id": q["id"], "category": q["category"],
                           "question": q["question"], "status": "ERROR", "secs": 0.0,
                           "answer": "", "error": str(exc)[:300]}
                f.write(json.dumps(rec, ensure_ascii=False) + "\n")
                f.flush()
                rows.append(rec)
                n_run += 1
                el = (time.time() - t0) / 60
                eta = el / n_run * (todo_total - n_run)
                print(f"[{n_run}/{todo_total}] {m} {q['id']} {rec['status']} "
                      f"({rec['secs']:.0f}s) | đã chạy {el:.0f}p, còn ~{eta:.0f}p", flush=True)
                time.sleep(args.delay)
        f.close()

    if not rows:
        print("Chưa có kết quả nào để dựng báo cáo.")
        return
    build_excel(rows, data["questions"], Path(args.out))
    print(f"\n✓ Excel: {args.out}")
    print(f"  Raw  : {results_path}")


if __name__ == "__main__":
    main()
